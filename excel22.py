# -*- coding: utf-8 -*-
#import xlrd
#import xlwt
import openpyxl
import urllib3
import json
import os


#def read_excel():
    # 打开文件
    #workbook = xlrd.open_workbook(r'.\file.xls')
    # 获取所有sheet
    #print (workbook.sheet_names())  # [u'sheet1', u'sheet2']
    #sheet2_name = workbook.sheet_names()[1]

    # 根据sheet索引或者名称获取sheet内容
    #sheet2 = workbook.sheet_by_index(1)  # sheet索引从0开始
    #sheet2 = workbook.sheet_by_name('Sheet1')

    # sheet的名称，行数，列数
    #print (sheet2.name, sheet2.nrows, sheet2.ncols)

    # 获取整行和整列的值（数组）
    #rows = sheet2.row_values(3)  # 获取第四行内容
    #cols = sheet2.col_values(0)# 获取第一列内容

f = openpyxl.load_workbook('file.xlsx')
sheet1 = f['Sheet1']
urllib3.disable_warnings()
http = urllib3.PoolManager(num_pools=100,maxsize=100, block=False)
#file = openpyxl.Workbook()
#table = file.create_sheet()
for i in range(1,190):
    #print(sheet1.cell(row = i ,column = 1).value)
    num=sheet1.cell(row = i ,column = 1).value
    #print(num)
    hao=str(num)
    #table.cell(row=i, column=1, value=hao)
    #urllib3.disable_warnings()
    #http = urllib3.PoolManager(num_pools=100,maxsize=100, block=False)
    url = 'https://cx.shouji.360.cn/phonearea.php?number=' + hao
    r = http.request('GET', url).data
    msg = json.loads(r.decode('utf-8'))
    #print(msg)
    #print(msg["data"])
    if len(msg["data"]["city"]) == 0:
        msg["data"]["city"] = msg["data"]["province"]

    value_list = list(msg["data"].values())
    #print(msg["data"])
    #guishudi = str(value_list)
    file = open("data.txt", "a")
    print(i, ":", hao, ":", value_list[1])
    print(i, ":", hao, ":", value_list[1], file=file)
    # table.cell(row=i,column=1,value=hao)
    #table.cell(row=i,column=2,value=guishudi)
    #file.save('data.xlsx')
    file.close()








#for i in cols:
        #print(int(i))
 #       num = int(i)
 #       hao = str(num)
 #       #print (num)
 #       urllib3.disable_warnings()
 #       http = urllib3.PoolManager()
 #       url = 'https://cx.shouji.360.cn/phonearea.php?number=' + hao
 #       r = http.request('GET', url)
 #       msg = json.loads(r.data.decode('utf-8'))
        #print(msg)
        #print(msg["data"])
 #       value_list = list(msg["data"].values())
 #       print(hao,str(value_list))

 #       file = openpyxl.Workbook()
 #       table = file.create_sheet('data')
 #       for n in range(10):
 #          table.cell(row=n+1,column=2,value=str(value_list));
 #          table.cell(row=n+1,column=1,value=hao)
 #       file.save('Excel_Workbook.xls')

    #print (cols)
    #print (sheet2.cell(9,0).ctype)

#if __name__ == '__main__':
#    read_excel()

os.system("pause")
